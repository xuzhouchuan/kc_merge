# -*- coding: utf-8 -*-
#!/usr/bin/env python
########################################################################
# 
# Copyright (c) 2018 Baidu.com, Inc. All Rights Reserved
# 
########################################################################
 
'''
File: merge_excels.py
Author: baidu(baidu@baidu.com)
Date: 2018/02/02 11:17:24
'''
import openpyxl as pxl
import xlrd
import xlwt
from xlutils.copy import copy
from xlutils.filter import process, XLRDReader, XLWTWriter
import sys
import os
import re
import copy
from optparse import OptionParser

def division(first, second):
    if second != 0:
        return 1.0 * first / second
    elif first == 0:
        return '--'

def get_keys(dict_data, key_str):
    keys = key_str.strip().split(':')
    d = dict_data
    for k in keys:
        d = d[k] 
    return d

def cell_name_to_coordinates(cell_name):
    col_name = [x for x in cell_name if x > '9' or x < '0'] 
    lin_name = [x for x in cell_name if x >= '0' and x <= '9']
    col = 0
    col_name.reverse()
    pow = 1
    for idx, d in enumerate(col_name):
        col += (int(d, 36) - 9) * pow
        pow *= 26
    col -= 1

    lin = int(''.join(lin_name)) - 1
    return (lin, col)

def get_next_cell_name(cur_cell, line_or_column=False, step=1): 
    reobj = re.match(r'([A-Z]+)(\d+)', cur_cell)
    alphabets = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    if reobj is None:
        return None
    lin_num = reobj.group(2)
    col_alpha = reobj.group(1)
    if not line_or_column:
        while step > 0:
            new_col = []
            up = 0
            i = len(col_alpha) - 1
            n = int(col_alpha[i], 36) - 10
            new_n = (n + 1 + up) % 26
            new_col.insert(0, alphabets[new_n]) 
            up = (1 if (n + 1 + up) > 25 else 0)
            i -= 1
            while i >= 0:
                n = int(col_alpha[i], 36) - 10
                new_n = (n + up) % 26 
                new_col.insert(0, alphabets[new_n]) 
                up = (1 if (n + up) > 25 else 0)
                i -= 1
            if up != 0:
                new_col.insert(0, alphabets[0])
            col_alpha = ''.join(new_col)
            step -= 1
        return ''.join(new_col) + lin_num
    else:
        lin_num = str(int(lin_num, 10) + step)
        return col_alpha + lin_num

busi_names = [u'测绘地理信息', u'管线工程', u'应用地球物理工程', u'新兴业务航空遥感', u'新兴业务智慧城市', u'其他新兴业务\n（测绘地理信息类）',
        u'其他新兴业务\n（管线工程类）', u'其他新兴业务\n（应用地球物理工程类）']


class ExcelModifier(object):
    def __init__(self):
        self.out_book = None
        self.style_list = None
        self.in_book = None

    def open_workbook(self, filename):
        wb  = xlrd.open_workbook(filename, formatting_info=True, on_demand=True, encoding_override='utf-8')
        w = XLWTWriter()
        process(XLRDReader(wb, "unknown.xls"), w)
        self.in_book = wb
        self.out_book = w.output[0][1]
        self.style_list = w.style_list
        return self

    def modify(self, sheet, x, y=-1, value=None, font=None):
        if (y < 0):
            x, y = cell_name_to_coordinates(x)
        if font is None:
            try:
                font = self._get_cell_style(sheet, x, y)
            except:
                pass
        if font is not None:
            self.out_book.get_sheet(sheet).write(x, y, value, font)
        else:
            self.out_book.get_sheet(sheet).write(x, y, value)

    def merge(self, sheet, from_x, from_y, to_x, to_y):
        self.out_book.get_sheet(sheet).merge(from_x, to_x, from_y, to_y)

    def merge_modify(self, sheet, from_x, from_y, to_x, to_y, value=None, font=None):
        if font is not None:
            self.out_book.get_sheet(sheet).write_merge(from_x, to_x, from_y, to_y, value, font)
        else:
            self.out_book.get_sheet(sheet).write_merge(from_x, to_x, from_y, to_y, value)
    def get_value(self, sheet, x, y):
        if (y < 0):
            x, y = cell_name_to_coordinates(x)
        return self.in_book.sheet_by_index(sheet).cell_value(x, y)

    def save(self, filename):
        self.out_book.save(filename)

    def _get_cell_style(self, sheet, x, y):
        xf_index = self.in_book.sheet_by_index(sheet).cell_xf_index(x, y)
        return self.style_list[xf_index]

class ExcelWriter(object):
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = pxl.Workbook()
        self.first_sheet = True
        self.current_sheet = self.wb.active
    
    def save(self):
        self.wb.save(self.file_name)

    def get_new_sheet(self, name):
        if self.first_sheet:
            self.first_sheet = False
            self.wb.active.title = name
            return self.wb.active
        else:
            ws = self.wb.create_sheet(name)
            self.current_sheet = ws
            return ws
        

    def style_range(self, cell_range, border=None, fill=None, font=None, alignment=None):
        ws = self.current_sheet
        first_cell = ws[cell_range.split(":")[0]]
        rows = ws[cell_range]
        if border is None:
            bd = pxl.styles.Side(style='thin', color='000000')
            border = pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
        for row in rows:
            for c in row:
                c.border = border
                if fill is not None:
                    c.fill = fill
                if font is not None:
                    c.font = font
                if c.alignment is not None:
                    c.alignment=alignment

    def fit_width(self, start_row=1):
        dims = {}
        idx = 0
        for row in self.current_sheet.rows:
            idx += 1
            if idx < start_row:
                continue
            for cell in row:
                #if cell.coordinate in self.current_sheet.merged_cells:
                #    continue
                if cell.value:
                    length = 0
                    if type(cell.value) == float:
                        length = 8
                    elif len(unicode(cell.value).encode('utf-8')) > 2 * len(unicode(cell.value)):
                        length = 2 * len(unicode(cell.value))
                    else:
                        length = len(unicode(cell.value))
                    dims[cell.column] = max(dims.get(cell.column, 0), length)

        for col, value in dims.items():
            self.current_sheet.column_dimensions[col].width = value * 1.1

class ExcelReader(object):
    def __init__(self):
        pass

class ExcelReader07(ExcelReader):
    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = pxl.load_workbook(file_name, data_only=True)
        self.current_sheet = None
    def set_current_sheet(self, sheet_name):
        try:
            self.current_sheet = self.wb[sheet_name]
            return True
        except Exception as e:
            return False

    def get_current_sheet_value(self, x, y=-1, default=None, type=None):
        if self.current_sheet is None:
            print u"file:%s, current_sheet is None" % self.file_name
            return None
        if (y < 0):
            ret = self.current_sheet[x].value
        else:
            x = x + 1
            y = y + 1
            ret = self.current_sheet.cell(x,y).value
        if (ret == '' or ret is None) and default is not None:
            ret = default
        if type:
            return type(ret)
        else:
            return ret


    def get_value(self, sheet, x, y=-1, default=None, type=None):
        if (y < 0):
            ret = self.wb.worksheets[sheet][x].value
            if (ret == '' or ret is None) and default is not None:
                ret = default
        else:
            x = x + 1
            y = y + 1
            ret = self.wb.worksheets[sheet].cell(x, y).value
            if (ret == '' or ret is None) and default is not None:
                ret = default
        if type:
            return type(ret)
        else:
            return ret

class ExcelReader03(ExcelReader):
    def __init__(self, file_name):
        self.in_book = xlrd.open_workbook(file_name, encoding_override='utf-8')

        self.current_sheet = None
        self.file_name = file_name
    def set_current_sheet(self, sheet_name):
        try:
            self.current_sheet = self.in_book.sheet_by_name(sheet_name)
            return True
        except Exception as e:
            return False

    def get_current_sheet_value(self, x, y=-1, default=None, type=None):
        if self.current_sheet is None:
            print u"file:%s, current_sheet is None" % self.file_name
            return None
        if (y < 0):
            x, y = cell_name_to_coordinates(x)

        ret = self.current_sheet.cell(x,y).value
        if (ret == '' or ret is None) and default is not None:
            ret = default
        if type:
            return type(ret)
        return ret

    def get_value(self, sheet, x, y=-1, default=None, type=None):
        if (y < 0):
            x, y = cell_name_to_coordinates(x)
        if y >= self.in_book.sheet_by_index(sheet).ncols or x >= self.in_book.sheet_by_index(sheet).nrows:
            return None

        ret = self.in_book.sheet_by_index(sheet).cell(x, y).value
        if (ret == '' or ret is None) and default is not None:
            ret = default
        if type:
            return type(ret)
        else:
            return ret


class ExcelMerger(object):
    def __init__(self):
        self.strategy_list = []

    def add_merge_strategy(self, merge_function):
        self.strategy_list.append(merge_function)

    def merge(self, to_excel, from_excel_list, from_key_list):
        for s in self.strategy_list:
            s.merge(to_excel, from_excel_list, from_key_list)

class SheetMergeFunction(object):
    def __init__(self):
        pass

    def merge(self, to_excel, from_excel):
        pass

class PersonSheetMergeFunction(SheetMergeFunction):
    def __init__(self):
        self.title = u'各单位市场部人员情况统计'
        self.heads = [ u'序号', u'单位名称', u'月初市场人员数', u'月末市场人员数', u'离职率', u'人员出勤率', u'人员培训状况', u'备注']
    def merge(self, to_excel, from_excel_list, from_keys):
        out_ws = to_excel.get_new_sheet(self.title) 
        out_ws['A1'] = self.title
        out_ws.merge_cells('A1:H1')
        out_ws['A1'].font = pxl.styles.Font(name=u'宋体',\
                size=16, bold=True)
        out_ws['A1'].alignment = pxl.styles.Alignment(horizontal='center')

        format_lines = len(from_keys) + 2
        format_range = 'A2:H%d' % (format_lines + 1)
        row_idx = len(from_keys) + 3
        out_ws.merge_cells('A%d:B%d' % (row_idx, row_idx))
        bd = pxl.styles.Side(style='thin', color='000000')
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=14), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))
        row = out_ws['A2:H2'][0]
        for i, c in enumerate(row):
            c.value = self.heads[i]

        sum_begin = 0
        sum_end = 0

        for idx, key in enumerate(from_keys):
            from_excel = from_excel_list[idx]
            row_idx = 3 + idx
            out_ws['A%d' % row_idx] = (idx + 1)
            out_ws['B%d' % row_idx] = key
            x = -1
            value = None
            for x_x in range(13, 1, -1):
                v = from_excel.get_value(0, x_x, 1)
                vv = from_excel.get_value(0, x_x, 2)
                v_str = v
                vv_str = vv
                if type(v) == float or type(v) == int or type(v) == long:
                    v_str = str(v)
                if type(vv) == float or type(vv) == int or type(vv) == long:
                    vv_str = str(vv)
                if v is not None and len(v_str) > 0 and vv is not None and len(vv_str) > 0:
                    value = v
                    x = x_x
                    break
            if value is None:
                print u"%s has no market-department people information" % key
                return False
            b_n = from_excel.get_value(0, x, 1, 0, int)
            e_n = from_excel.get_value(0, x, 2, 0, int)
            out_ws['C%d' % row_idx] = b_n
            out_ws['D%d' % row_idx] = e_n
            out_ws['E%d' % row_idx] = division(b_n - e_n, b_n)
            out_ws['E%d' % row_idx].number_format = '0.00%'
            out_ws['F%d' % row_idx] = 1
            out_ws['F%d' % row_idx].number_format = '0.00%'
            #out_ws['F%d' % row_idx].guess_types = True
            out_ws['G%d' % row_idx] = u'无'
            out_ws['H%d' % row_idx] = from_excel.get_value(0, x, 4, '')
            sum_begin += from_excel.get_value(0, x, 1)
            sum_end += int(from_excel.get_value(0, x, 2, 0))
        row_idx = len(from_keys) + 3
        out_ws['A%d' % row_idx] = u'合计'
        out_ws['C%d' % row_idx] = sum_begin
        out_ws['D%d' % row_idx] = sum_end
        out_ws['C%d' % row_idx].font= pxl.styles.Font(size=15, color='ff0000')
        out_ws['D%d' % row_idx].font= pxl.styles.Font(size=15, color='ff0000')
        to_excel.fit_width(2)
        return True

class ReturnMoneyMergeFunction(SheetMergeFunction):
    def __init__(self):
        self.from_title = u'附表6 本月回款情况统计表'
        self.sheet_title = u'回款情况'
        self.title1 = u'各单位回款情况'
        self.title2 = u'各板块回款情况'
        self.heads = [u'序号', u'名称', u'回款总额（万元）', u'回款占比']

    def merge(self, to_excel, from_excel_list, from_keys):
        #summary
        company_busi_return_money_map = {}
        for idx, key in enumerate(from_keys):
            company_busi_return_money_map[key] = {}
            from_excel = from_excel_list[idx]
            set_sheet_ret = from_excel.set_current_sheet(self.from_title) 
            if not set_sheet_ret:
                title = u'附表6本月回款情况统计表'
                set_sheet_ret = from_excel.set_current_sheet(title)
                if not set_sheet_ret:
                    print key
                    print "no sheet named:%s" % self.from_title
                    sys.exit(1)
            row_i = 4
            total_money = 0.0
            while True:
                try:
                    value = from_excel.get_current_sheet_value('E%d' % row_i, default=-1.0, type=float)
                except:
                    break
                value = value / 10000.0
                busi_name = from_excel.get_current_sheet_value('D%d' % row_i)
                if value < 0.0:
                    break 
                is_total = from_excel.get_current_sheet_value('A%d' % row_i)
                if is_total == u'合计':
                    break
                if busi_name is None or busi_name == '':
                    busi_name = "未知分类"
                busi_name = busi_name.strip()
                total_money += value
                if company_busi_return_money_map[key].get(busi_name) is None:
                    company_busi_return_money_map[key][busi_name] = 0.0
                company_busi_return_money_map[key][busi_name] += value
                row_i += 1
            company_busi_return_money_map[key]['total'] = total_money

        all_total_money = 0.0
        busi_return_total_money_map = {}
        for k, x in company_busi_return_money_map.iteritems():
            for t, y in x.iteritems():
                if t == 'total':
                    all_total_money += y
                else:
                    if t not in busi_return_total_money_map:
                        busi_return_total_money_map[t] = 0.0
                    busi_return_total_money_map[t] += y
        

        #write values
        out_ws = to_excel.get_new_sheet(self.sheet_title)
        out_ws['A1'] = self.title1
        out_ws.merge_cells('A1:D1')
        #title style
        out_ws['A1'].font = pxl.styles.Font(name=u'宋体',\
                size=16, bold=True)
        out_ws['A1'].alignment = pxl.styles.Alignment(horizontal='center')
        #TODO format content
        row = out_ws['A2:D2'][0]
        #write table heads
        for i, c in enumerate(row):
            c.value = self.heads[i]
        idx = 1 

        range_start = 'A2'
        range_end = 'D2'

        for com_name, value in company_busi_return_money_map.iteritems():
            cell_idx = idx + 2
            out_ws['A%d' % cell_idx] = idx
            out_ws['B%d' % cell_idx] = com_name
            out_ws['C%d' % cell_idx] = value['total']
            out_ws['D%d' % cell_idx] = value['total'] / all_total_money
            out_ws['D%d' % cell_idx].number_format = '0.00%'
            idx += 1
        #总计
        cell_idx = idx + 3
        out_ws['B%d' % cell_idx] = '总计'
        out_ws['C%d' % cell_idx] = all_total_money
        range_end = 'D%d' % cell_idx
        
        format_range = '%s:%s' % (range_start, range_end)
        bd = pxl.styles.Side(style='thin', color='000000')
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=12), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))


        cell_idx = idx + 6
        out_ws['A%d' % cell_idx] = self.title2
        out_ws.merge_cells('A%d:D%d' % (cell_idx, cell_idx))
        #title style
        out_ws['A%d' % cell_idx].font = pxl.styles.Font(name=u'宋体',\
                size=16, bold=True)
        out_ws['A%d' % cell_idx].alignment = pxl.styles.Alignment(horizontal='center')
        #TODO format content
        cell_idx += 1
        range_start = 'A%d' % cell_idx
        range_end = 'D%d' % cell_idx
        row = out_ws['A%d:D%d' % (cell_idx, cell_idx)][0]
        #write table heads
        for i, c in enumerate(row):
            c.value = self.heads[i]

        idx = 1
        row_idx = 0
        for busi_name in busi_names:
            busi_value = busi_return_total_money_map.get(busi_name, 0.0)
            row_idx = cell_idx + idx
            out_ws['A%d' % row_idx] = idx
            out_ws['B%d' % row_idx] = busi_name
            out_ws['C%d' % row_idx] = busi_value
            out_ws['D%d' % row_idx] = busi_value / all_total_money
            out_ws['D%d' % row_idx].number_format = '0.00%'
            idx += 1

        row_idx = cell_idx + idx + 1
        out_ws['B%d' % row_idx] = '总计'
        out_ws['C%d' % row_idx] = all_total_money
        range_end = 'D%d' % row_idx
        row_idx += 1 
        out_ws['A%d' % row_idx] = 'exception:'
        for com_name, v in company_busi_return_money_map.iteritems():
            for t, y in v.iteritems():
                if t != u'total' and t not in busi_names:
                    row_idx += 1
                    out_ws['A%d' % row_idx] = com_name
                    out_ws['B%d' % row_idx] = t
                    out_ws['C%d' % row_idx] = y


        format_range = '%s:%s' % (range_start, range_end)
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=12), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))



    #out_ws['A%d' % (idx + 3)] = (idx + 1)
            #out_ws['B%d' % (idx + 3)] = key





class BiddingSheetMergeFunction(SheetMergeFunction):
    def __init__(self):
        self.title = u'投标汇报数据'
    def merge(self, to_excel, from_excel_list, from_keys):
        self._region_statistic(to_excel, from_excel_list, from_keys) 

    def _region_statistic(self, to_excel, from_excel_list, from_keys):
        bidding_map = {}
        max_region_num = 0
        for idx, key in enumerate(from_keys):
            from_excel = from_excel_list[idx]
            start_x = -1
            start_y = 3
            for i in range(0, 100):
                v = from_excel.get_value(3, i, 2)
                if v == u'市场区域':
                    start_x = i
                    break
            bidding_info = []
            for i in range(start_y, 50):
                region = from_excel.get_value(3, start_x, i)
                print region
                if region is not None and len(unicode(region)) > 0:
                    if region.endswith(u'省'):
                        region = region[:-1]
                    elif region.endswith(u'自治区'):
                        region = region[:-3]
                    elif region.endswith(u'特别行政区'):
                        region = region[:-5]
                    bidding_num = from_excel.get_value(3, start_x + 1, i, 0)
                    hit_num = from_excel.get_value(3, start_x + 2, i, 0)
                    hit_ratio = from_excel.get_value(3, start_x + 3, i, 0)
                    bidding_info.append((region, bidding_num, hit_num, hit_ratio))
                    bidding_map[key] = bidding_info
                else:
                    break
            if key in bidding_map and len(bidding_map[key]) > max_region_num:
                max_region_num = len(bidding_map[key])
        #write
        max_width = max_region_num + 6
        out_ws = to_excel.get_new_sheet(self.title)
        out_ws['A1'] = self.title
        out_ws['A1'].font = pxl.styles.Font(name=u'宋体', size=16, bold=True)
        out_ws['A1'].alignment = pxl.styles.Alignment(horizontal='center')
        out_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_width)
        max_col_name = out_ws.cell(1, max_width).column
        format_range = 'A2:%s%d' % (max_col_name, len(from_keys) * 4 + 1)
        bd = pxl.styles.Side(style='thin', color='000000')
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=14), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))

        total_info = {}
        region_info = {}
        t_bidding = 0
        h_bidding = 0
        for idx, key in enumerate(from_keys):
            start_x = 4 * idx + 2
            end_x = 4 * idx + 5
            out_ws.merge_cells('A%d:A%d' % (start_x, end_x))
            out_ws['A%d' % start_x] = (idx + 1)
            out_ws.merge_cells('B%d:B%d' % (start_x, end_x))
            out_ws['B%d' % start_x] = key
            out_ws.merge_cells('C%d:C%d' % (start_x, end_x))
            out_ws['C%d' % start_x] = u'招投标情况'
            out_ws['B%d' % start_x].fill = pxl.styles.PatternFill('solid', fgColor='ffff00')
            out_ws['D%d' % start_x] = u'市场区域'
            out_ws['D%d' % (start_x + 1)] = u'投标次数'
            out_ws['D%d' % (start_x + 2)] = u'中标次数'
            out_ws['D%d' % (start_x + 3)] = u'中标率'
            out_ws['E%d' % start_x] = u'市场区域'
            out_ws['E%d' % (start_x + 1)] = u'投标次数'
            out_ws['E%d' % (start_x + 2)] = u'中标次数'
            out_ws['E%d' % (start_x + 3)] = u'中标率'
            out_ws['F%d' % start_x] = u'合计'
            total_bidding = 0
            hit_bidding = 0
            province_num = 0
            if key in bidding_map:
                for j, info in enumerate(bidding_map[key]):
                    if info[1] > 0:
                        province_num += 1 
                    total_bidding += int(info[1])
                    hit_bidding += int(info[2])
                    if info[0] not in region_info:
                        region_info[info[0]] = [0, 0]
                    region_info[info[0]][0] += int(info[1])
                    region_info[info[0]][1] += int(info[2])
                    for i in range(4):
                        out_ws.cell(start_x + i, 7 + j).value = info[i]
                    out_ws.cell(start_x + 3, 7 + j).number_format = '0.00%'
            total_info[key] = (idx, key, total_bidding, hit_bidding, division(hit_bidding, total_bidding), province_num)
            out_ws['F%d' % (start_x + 1)] = total_bidding
            out_ws['F%d' % (start_x + 2)] = hit_bidding
            out_ws['F%d' % (start_x + 3)] = division(hit_bidding, total_bidding)
            out_ws['F%d' % (start_x + 3)].number_format = '0.00%'
            h_bidding += hit_bidding
            t_bidding += total_bidding
            
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=12), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))
        #---------------------------------------------------------------------------------
        start_x = len(from_keys) * 4 + 3
        format_range = 'B%d:G%d' % (start_x, start_x + len(from_keys) + 1)
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=12), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))
        titles = [u'序号', u'公司名称', u'投标次数', u'中标次数', u'中标率', u'投标省份(个)']
        for idx, t in enumerate(titles):
            out_ws.cell(start_x, 2 + idx, t)

        for idx, key in enumerate(from_keys):
            for j in range(6):
                
                out_ws.cell(start_x + 1 + idx, j + 2, total_info[key][j])

        start_x += len(from_keys) + 1
        out_ws['C%d' % start_x] = u'合计'
        out_ws['D%d' % start_x] = t_bidding
        out_ws['E%d' % start_x] = h_bidding
        out_ws['F%d' % start_x] = 1.0 * h_bidding / t_bidding
        out_ws['F%d' % start_x].number_format = '0.00%'
        
        start_x += 2
        titles = [u'序号', u'投标区域', u'投标次数', u'中标次数', u'中标率']
        
        #---------------------------------------------------------------------
        format_range = 'B%d:F%d' % (start_x, start_x + len(region_info))
        to_excel.style_range(cell_range=format_range, \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                font=pxl.styles.Font(name=u'宋体', size=12), \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))

        for idx, t in enumerate(titles):
            out_ws.cell(start_x, idx + 2, t)
       
        idx = 0
        for key, info in region_info.iteritems():
            idx += 1
            out_ws['B%d' % (start_x + idx)] = idx
            out_ws['C%d' % (start_x + idx)] = key
            out_ws['D%d' % (start_x + idx)] = info[0]
            out_ws['E%d' % (start_x + idx)] = info[1]
            out_ws['F%d' % (start_x + idx)] = division(info[1], info[0])
            out_ws['F%d' % (start_x + idx)].number_format = '0.00%'


        to_excel.fit_width(2)

class ContractSheetMergeFunction(SheetMergeFunction):
    def __init__(self):
        self.title = u'合同分析表'

    def merge(self, to_excel, from_excel_list, from_keys):
        #statistic
        contract_num = {}
        contract_amount = {}
        contract_detail = {}
        for idx, key in enumerate(from_keys):
            from_excel = from_excel_list[idx]
            contract_num[key] = {"this_month": {}, "last_month": {}, "smly": {}}
            contract_amount[key] = {"this_month": {}, "last_month": {}, "smly": {}}
            contract_detail[key] = {}
            smly_num = contract_num[key]["smly"]
            smly_num["sum"] = from_excel.get_value(2, "B7", -1, 0, int)
            smly_num["carry"] = from_excel.get_value(2, "C7", -1, 0, int)
            smly_num["new"] = from_excel.get_value(2, "E7", -1, 0, int)
            smly_num["accum"] = from_excel.get_value(2, "D7", -1, 0, int)

            thism_num = contract_num[key]["this_month"]
            thism_num["sum"] = from_excel.get_value(2, "F7", -1, 0, int)
            thism_num["carry"] = from_excel.get_value(2, "G7", -1, 0, int)
            thism_num["accum"] = from_excel.get_value(2, "H7", -1, 0, int)
            thism_num["new"] = from_excel.get_value(2, "I7", -1, 0, int)

            smly_amount = contract_amount[key]["smly"]
            smly_amount["sum"] = from_excel.get_value(2, "J7", -1, 0.0, float)
            smly_amount["carry"] = from_excel.get_value(2, "K7", -1, 0.0, float)
            smly_amount["accum"] = from_excel.get_value(2, "L7", -1, 0.0, float)
            smly_amount["new"] = from_excel.get_value(2, "M7", -1, 0.0, float)

            thism_amount = contract_amount[key]["this_month"]
            thism_amount["sum"] = from_excel.get_value(2, "N7", -1, 0.0, float)
            thism_amount["carry"] = from_excel.get_value(2, "O7", -1, 0.0, float)
            thism_amount["accum"] = from_excel.get_value(2, "P7", -1, 0.0, float)
            thism_amount["new"] = from_excel.get_value(2, "Q7", -1, 0.0, float)
            
            #details
            details = contract_detail[key]
            for d_idx in range(8):
                lin_name = d_idx + 8
                busi_name = from_excel.get_value(2, "A%d" % (lin_name), -1)
                details[busi_name] = {"num" : {"this_month": {}, "last_month": {}, "smly": {}},
                        "amount" : {"this_month": {}, "last_month": {}, "smly": {}}}
                num = details[busi_name]["num"]
                num["smly"]["sum"] = from_excel.get_value(2, "B%d" % lin_name, -1, 0.0, float)
                num["smly"]["carry"] = from_excel.get_value(2, "C%d" % lin_name, -1, 0.0, float)
                num["smly"]["accum"] = from_excel.get_value(2, "D%d" % lin_name, -1, 0.0, float)
                num["smly"]["new"] = from_excel.get_value(2, "E%d" % lin_name, -1, 0.0, float)

                num["this_month"]["sum"] = from_excel.get_value(2, "F%d" % lin_name, -1, 0.0, float)
                num["this_month"]["carry"] = from_excel.get_value(2, "G%d" % lin_name, -1, 0.0, float)
                num["this_month"]["accum"] = from_excel.get_value(2, "H%d" % lin_name, -1, 0.0, float)
                num["this_month"]["new"] = from_excel.get_value(2, "I%d" % lin_name, -1, 0.0, float)

                amount = details[busi_name]["amount"]
                amount["smly"]["sum"] = from_excel.get_value(2, "J%d" % lin_name, -1, 0.0, float)
                amount["smly"]["carry"] = from_excel.get_value(2, "K%d" % lin_name, -1, 0.0, float)
                amount["smly"]["accum"] = from_excel.get_value(2, "L%d" % lin_name, -1, 0.0, float)
                amount["smly"]["new"] = from_excel.get_value(2, "M%d" % lin_name, -1, 0, float)

                amount["this_month"]["sum"] = from_excel.get_value(2, "N%d" % lin_name, -1, 0.0, float)
                amount["this_month"]["carry"] = from_excel.get_value(2, "O%d" % lin_name, -1, 0.0, float)
                amount["this_month"]["accum"] = from_excel.get_value(2, "P%d" % lin_name, -1, 0.0, float)
                amount["this_month"]["new"] = from_excel.get_value(2, "Q%d" % lin_name, -1, 0.0, float)
        out_ws = to_excel.get_new_sheet(self.title)
        out_ws['A1'].font = pxl.styles.Font(name=u'宋体', size=16, bold=True) 
        out_ws['A1'] = self.title + u'(单位：万元)'
        out_ws.merge_cells('A1:K1')
        titles = [ u'单位名称',
                u'本月新签合同额',
                u'本月累计新签合同额',
                u'本月合同总额',
                u'去年同期新签合同额',
                u'去年同期累计新签合同额',
                u'去年同期合同总额',
                u'本月累计合同份数',
                u'本月新签合同份数',
                u'本月合同总份数',
                u'去年同期累计合同份数',
                u'去年同期新签合同份数',
                u'去年同期合同总份数']

        cell_name = 'A2'
        for i, t in enumerate(titles):
            out_ws[cell_name] = t
            out_ws.merge_cells('%s:%s' % (cell_name, get_next_cell_name(cell_name, True)))
            cell_name = get_next_cell_name(cell_name)
        
        start_cell_name = 'A4'
        format_start = 'A1'
        new_sum_amount = 0
        for key in from_keys:
            new_sum_amount += contract_amount[key]['this_month']['new']
        for key in from_keys:
            out_ws[start_cell_name] = key
            cur_cell = get_next_cell_name(start_cell_name)
            out_ws[cur_cell] = contract_amount[key]['this_month']['new']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_amount[key]['this_month']['accum']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_amount[key]['this_month']['sum']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_amount[key]['smly']['new']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_amount[key]['smly']['accum']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_amount[key]['smly']['sum']

            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_num[key]['this_month']['accum']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_num[key]['this_month']['new']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_num[key]['this_month']['sum']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_num[key]['smly']['accum']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_num[key]['smly']['new']
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = contract_num[key]['smly']['sum']

            format_end = cur_cell
            start_cell_name = get_next_cell_name(start_cell_name, True)

        bd = pxl.styles.Side(style='thin', color='000000')
        to_excel.style_range(cell_range="%s:%s" % (format_start, format_end), \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=None, \
                alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))
        
        start_cell_name = get_next_cell_name(start_cell_name, True)
        start_cell_name = get_next_cell_name(start_cell_name, True)
        out_ws[start_cell_name] = u'本月各公司各类项目合同额'
        out_ws[start_cell_name].font = pxl.styles.Font(name=u'宋体', size=16, bold=True)
        format_start = start_cell_name
        busi_names = [u'测绘地理信息', u'管线工程', u'应用地球物理工程', u'新兴业务航空遥感', u'新兴业务智慧城市', u'其他新兴业务\n（测绘地理信息类）',
                u'其他新兴业务\n（管线工程类）', u'其他新兴业务\n（应用地球物理工程类）']
        titles = copy.deepcopy(busi_names)
        titles.insert(0, u'各单位名称')
        titles.extend([u'新兴业务新签合同', u'备注'])
        out_ws.merge_cells('%s:%s' % (start_cell_name, get_next_cell_name(start_cell_name, False, len(titles) + len(busi_names) - 1)))
        start_cell_name = get_next_cell_name(start_cell_name, True)
        cell_name = start_cell_name
        format_end =start_cell_name
        for i, t in enumerate(titles):
            out_ws[cell_name] = t
            cell_end = cell_name
            if t in busi_names:
                cell_end = get_next_cell_name(get_next_cell_name(cell_name))
            else:
                cell_end = get_next_cell_name(cell_name, True)
            out_ws.merge_cells('%s:%s' % (cell_name, cell_end))
            if t in busi_names:
                first_cell = get_next_cell_name(cell_name, True)
                second_cell = get_next_cell_name(get_next_cell_name(cell_name, True))
                third_cell = get_next_cell_name(get_next_cell_name(get_next_cell_name(cell_name, True)))
                out_ws[first_cell] = u'新增合同额'
                out_ws[second_cell] = u'今年累计合同额'
                out_ws[third_cell] = u'去年累计合同额'
                format_end = second_cell
                cell_name = get_next_cell_name(get_next_cell_name(get_next_cell_name(cell_name)))
            else:
                format_end = get_next_cell_name(cell_name, True)
                cell_name = get_next_cell_name(cell_name)

        bd = pxl.styles.Side(style='thin', color='000000')
        to_excel.style_range("%s:%s" % (start_cell_name, format_end), \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=pxl.styles.PatternFill('solid', fgColor='ffff00'), \
                font=pxl.styles.Font(name=u'宋体', size=14, bold=True))

        start_cell_name = get_next_cell_name(start_cell_name, True)
        start_cell_name = get_next_cell_name(start_cell_name, True)
        
        new_busi_info = {}
        for key in from_keys:
            out_ws[start_cell_name] = key
            cur_cell = get_next_cell_name(start_cell_name)
            new_busi_sum = 0.0
            for busi in busi_names:
                out_ws[cur_cell] = contract_detail[key][busi]['amount']['this_month']['new']
                cur_cell = get_next_cell_name(cur_cell)
                out_ws[cur_cell] = contract_detail[key][busi]['amount']['this_month']['accum']
                cur_cell = get_next_cell_name(cur_cell)
                out_ws[cur_cell] = contract_detail[key][busi]['amount']['smly']['accum']
                cur_cell = get_next_cell_name(cur_cell)
                if busi.find(u'新兴') >= 0:
                    new_busi_sum += contract_detail[key][busi]['amount']['this_month']['new']
                    if contract_detail[key][busi]['amount']['this_month']['new'] > 0.0:
                        if key not in new_busi_info:
                            new_busi_info[key] = []
                        new_busi_info[key].append(busi)
            out_ws[cur_cell] = new_busi_sum
            cur_cell = get_next_cell_name(cur_cell)
            if key in new_busi_info:
                out_ws[cur_cell] = u"、".join(new_busi_info[key])
            format_end = cur_cell 
            start_cell_name = get_next_cell_name(start_cell_name, True)
        to_excel.style_range(cell_range="%s:%s" % (format_start, format_end), alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))
        #every business
        busi_new_amount = [0.0] * len(busi_names)
        busi_accum_amount = [0.0] * len(busi_names)
        all_busi_accum_amount = 0.0
        for i, busi in enumerate(busi_names):
            for key in from_keys:
                busi_new_amount[i] += contract_detail[key][busi]['amount']['this_month']['new'] 
                busi_accum_amount[i] += contract_detail[key][busi]['amount']['this_month']['accum'] 
                all_busi_accum_amount += contract_detail[key][busi]['amount']['this_month']['accum'] 



        start_cell_name = get_next_cell_name(start_cell_name, True)
        single_tname = [u'新增合同额', u'今年累计合同额', u'去年累计合同额']
        single_key = ['amount:this_month:new', 'amount:this_month:accum', 'amount:smly:accum']
        for idx, tname in enumerate(single_tname):
            format_start = start_cell_name
            out_ws[start_cell_name] = u'按版块各数据单表:%s' % tname
            titles = [u'各单位名称',]
            titles.extend(busi_names)
            out_ws.merge_cells('%s:%s' % (start_cell_name, get_next_cell_name(start_cell_name, False, len(titles) -1)))
            start_cell_name = get_next_cell_name(start_cell_name, True)
            cell_name = start_cell_name
            for t in titles:
                out_ws[cell_name] = t
                cell_name = get_next_cell_name(cell_name)
            start_cell_name = get_next_cell_name(start_cell_name, True)
            for key in from_keys:
                out_ws[start_cell_name] = key
                cur_cell = get_next_cell_name(start_cell_name)
                for busi in busi_names:
                    out_ws[cur_cell] = get_keys(contract_detail[key][busi], single_key[idx])
                    format_end = cur_cell
                    cur_cell = get_next_cell_name(cur_cell)
                start_cell_name = get_next_cell_name(start_cell_name, True)

            to_excel.style_range(cell_range='%s:%s' % (format_start, format_end), \
                    border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                    fill=None, \
                    font=pxl.styles.Font(name=u'宋体', size=12), \
                    alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))

        
        start_cell_name = get_next_cell_name(start_cell_name, True)
        format_start = start_cell_name
        out_ws[start_cell_name] = u'各类项目本月合同额统计'
        out_ws[start_cell_name].font = pxl.styles.Font(name=u'宋体', size=16, bold=True)
        out_ws[start_cell_name].alignment = pxl.styles.Alignment(horizontal='center')
        titles = [u'各单位名称', u'新签合同额', u'新签合同额占比', u'累计合同额', u'累计合同额占比']
        out_ws.merge_cells('%s:%s' % (start_cell_name, get_next_cell_name(start_cell_name, False, len(titles) - 1)))

        start_cell_name = get_next_cell_name(start_cell_name, True)
        cur_cell = start_cell_name
        title_start = start_cell_name
        for t in titles:
            out_ws[cur_cell] = t
            cur_cell = get_next_cell_name(cur_cell)
        start_cell_name = get_next_cell_name(start_cell_name, True)

        to_excel.style_range("%s:%s" % (title_start, get_next_cell_name(title_start, False, len(titles) - 1)), \
                border=pxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd), \
                fill=pxl.styles.PatternFill('solid', fgColor='ffff00'), \
                font=pxl.styles.Font(name=u'宋体', size=14, bold=True))

        for i, busi in enumerate(busi_names):
            cur_cell = start_cell_name
            out_ws[cur_cell] = busi
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = busi_new_amount[i]
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = division(busi_new_amount[i], new_sum_amount)
            out_ws[cur_cell].number_format = '0.00%'
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = busi_accum_amount[i]
            cur_cell = get_next_cell_name(cur_cell)
            out_ws[cur_cell] = division(busi_accum_amount[i], all_busi_accum_amount)
            out_ws[cur_cell].number_format = '0.00%'
            format_end = cur_cell
            start_cell_name = get_next_cell_name(start_cell_name, True)

        to_excel.style_range(cell_range="%s:%s" % (format_start, format_end), alignment=pxl.styles.Alignment(horizontal='center', vertical='center'))

        to_excel.fit_width(2)


                    
def get_sub_excels(sub_dir):
    if not os.path.isdir(sub_dir):
        print u"sub_dir:%s is not a dir name" % sub_dir
    files = os.listdir(sub_dir)
    excel_list = []
    excel_key_list = []
    for f in files:
        ext = os.path.splitext(f)[1]
        if os.path.isfile(os.path.join(sub_dir, f)) and (ext == '.xls' or ext == '.xlsx'):
            excel_reader = None
            if ext == '.xls':
                excel_reader = ExcelReader03(os.path.join(sub_dir, f))
            else:
                excel_reader = ExcelReader07(os.path.join(sub_dir, f))
            excel_list.append(excel_reader)
            print f
            tmp_keys = os.path.splitext(f)[0].split('--')
            key = None
            if len(tmp_keys) == 2:
                key = tmp_keys[1]
            if key is None:
                tmp_keys = os.path.splitext(f)[0].split('-')
            if len(tmp_keys) == 2:
                key = tmp_keys[1]
            if key is None:
                tmp_keys = os.path.splitext(f)[0].split('——')
            if len(tmp_keys) == 2:
                key = tmp_keys[1]
            if key is None:
                print "file name error"
                sys.exit(1)

            excel_key_list.append(key)
    return (excel_list, excel_key_list)

def make_strategy():
    strategy_driver = ExcelMerger() 
    strategy_driver.add_merge_strategy(PersonSheetMergeFunction())
    strategy_driver.add_merge_strategy(BiddingSheetMergeFunction())
    strategy_driver.add_merge_strategy(ContractSheetMergeFunction())
    strategy_driver.add_merge_strategy(ReturnMoneyMergeFunction())
    return strategy_driver

def main(argv):
    parser = OptionParser()
    parser.add_option("-d", "--sub_dir", dest="sub_dir", metavar="DIR", default="subs", help="DIR: vector file path")
    parser.add_option("-t", "--template_file", dest="template", metavar="FILE", default="template.xls", help="FILE: output template excel file")
    parser.add_option("-o", "--output_file", dest="output_file", metavar="FILE", default="merge.xlsx", help="FILE: output file name")
    (options, args) = parser.parse_args(argv)

    in_excel_list, in_key_list = get_sub_excels(options.sub_dir)

    out_excel = ExcelWriter(options.output_file)

    strategy_driver = make_strategy()
    strategy_driver.merge(out_excel, in_excel_list, in_key_list)

    out_excel.save()




if __name__ == '__main__':
    main(sys.argv[1:])
